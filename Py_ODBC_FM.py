import pyodbc
import openpyxl
import configparser

def filemaker_odbc_connection(path):
    wb = openpyxl.load_workbook(path)
    ws = wb['Export']
    total_rows = ws.max_row

    config = configparser.ConfigParser()
    config.read('config.ini', encoding='utf-8')
    server = config['odbc']['server']
    database = config['odbc']['database']
    uid = config['odbc']['uid']
    pasd = config['odbc']['pasd']
    # ODBC 連接字串
    connection_string = f"Driver={{FileMaker ODBC}}; Server={server}; Database={database}; UID={uid}; PWD={pasd}"

    try:
        # 嘗試連接 FileMaker ODBC
        conn = pyodbc.connect(connection_string)
        cursor = conn.cursor()

        table_name = "人事_刷卡記錄"
        fields_to_insert = ["_pk_員工編號", "上班日期", "類別", "上班時間", "下班時間", "年", "月", "日", "lat", "lng", "ip", "大樓編號", "大樓名稱"]

        for i in range(2, total_rows + 1):
            data_dict = {
                "_pk_員工編號": ws.cell(i, 2).value,
                "上班日期": ws.cell(i, 5).value,
                "類別": ws.cell(i, 4).value,
                "上班時間": ws.cell(i, 6).value,
                "下班時間": ws.cell(i, 7).value,
                "年": int(ws.cell(i, 8).value),
                "月": int(ws.cell(i, 9).value),
                "日": int(ws.cell(i, 10).value),
                "lat": ws.cell(i, 12).value,
                "lng": ws.cell(i, 13).value,
                "ip": ws.cell(i, 17).value,
                "大樓編號": ws.cell(i, 15).value,
                "大樓名稱": ws.cell(i, 16).value,
            }

            # 使用 str.join() 方法構建 SQL INSERT 語句
            fields_str = ", ".join(f'"{field}"' for field in fields_to_insert)
            values_str = ", ".join("?" for _ in fields_to_insert)
            sql = f"INSERT INTO \"{table_name}\" ({fields_str}) VALUES ({values_str})"

            # 將字典中的值轉換為 list，並按照 fields_to_insert 順序排列
            data_values = [data_dict[field] for field in fields_to_insert]

            # 假設有 _pk_員工編號值，才執行 INSERT
            if data_dict["_pk_員工編號"]:
                cursor.execute(sql, data_values)
                conn.commit()

        cursor.close()
        conn.close()

    except Exception as e:
        print("連接失敗：", e)
