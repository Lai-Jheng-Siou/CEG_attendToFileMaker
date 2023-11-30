import openpyxl
from datetime import datetime, timedelta

'''
2: 員工編號, 3: 員工姓名, 4: 類別(上班or下班), 5: 日期
6: 上班時間, 7: 下班時間, 12: Lat, 13: Lng
'''
def detectAbsence(path, determine_time = "17:30"):
    xlsx = path
    wb = openpyxl.load_workbook(xlsx)
    ws = wb['Export']
    total_rows = ws.max_row + 1

    string = ""
    delta = timedelta(minutes = 30)
    count = 0  #計算如果只打卡一次 則判斷時間
    determineTime = datetime.strptime(determine_time, "%H:%M")  #以17:30判斷 如果大於則是下班 小於則是上班
    date = ws.cell(2, 5).value
    empId = ws.cell(2, 2).value
    name = ""
    onDuty = ""
    onDuty_pos = False
    offDuty = ""
    offDuty_pos = False

    for i in range(2, total_rows):
        next_id = ws.cell(i, 2).value
        if empId != next_id:
            if count == 1 and offDuty < determineTime:
                onDuty = offDuty
                onDuty_pos = offDuty_pos
                offDuty = ""
                offDuty_pos = ""
            elif count == 2 and onDuty != "" and offDuty != "" and offDuty - onDuty < delta:
                if offDuty >= determineTime:
                    onDuty = ""
                    onDuty_pos = ""
                else:
                    offDuty = ""
                    offDuty_pos = ""
            st = f"{date} {name} "
            if onDuty == "":
                string += st + "上班未打卡\n"
            elif not onDuty_pos:
                string += st + "上班未定位\n"
            if offDuty == "":
                string += st + "下班未打卡\n"
            elif not offDuty_pos:
                string += st + "下班未定位\n"
                
            empId = next_id
            name = ws.cell(i, 3).value
            onDuty = ""
            offDuty = ""
            onDuty_pos = False
            offDuty_pos = False
            count = 0
        
        onWork = ws.cell(i, 6).value
        offWork = ws.cell(i, 7).value
        lat = ws.cell(i, 12).value
        lng = ws.cell(i, 13).value

        str_time = onWork if onWork != None else offWork
        booling = True if lat != 0 and lng != 0 else False
        time = datetime.strptime(str_time, "%H:%M")
        
        if offDuty == "":
            offDuty = time
            offDuty_pos = booling
        else:
            last = offDuty + delta

            if time > last:  #如果下一個時間比上一個時間超出30分鐘 則上一個時間為上班時數 下一個則是下班時間
                onDuty = offDuty
                onDuty_pos = offDuty_pos
                offDuty = time
                offDuty_pos = True if offDuty_pos else booling
            else:  #反之更新定位就好
                offDuty_pos = True if offDuty_pos else booling

            if time < offDuty:  #如果下一個時間比上一個時間小 則下一個時間為上班時數
                onDuty = time
                onDuty_pos = True if onDuty_pos else booling

        count += 1

    if count == 1 and offDuty < determineTime:
        onDuty = offDuty
        onDuty_pos = offDuty_pos
        offDuty = ""
        offDuty_pos = ""
    st = f"{date} {name} "
    if onDuty == "":
        string += st + "上班未打卡\n"
    elif not onDuty_pos:
        string += st + "上班未定位\n"
    if offDuty == "":
        string += st + "下班未打卡\n"
    elif not offDuty_pos:
        string += st + "下班未定位\n"

    return string

if __name__ == '__main__':
    st = detectAbsence(r'C:\Users\CEG\Desktop\attend.xlsx')
    print(st)